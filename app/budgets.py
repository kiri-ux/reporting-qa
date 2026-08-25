"""Top up the money columns on order lines that are already loaded.

The nightly export did not carry a budget until it was added to the report in
the IO tool, and re-pulling the whole thing is a couple of million rows for the
sake of one column. So a sheet covering ONE product can be dropped in and it
fills the budget on the lines it names, leaving everything else alone.

MERGE, NEVER REPLACE. This is the important part. A file of 368 Performance Max
rows put through the normal import with replace=True would delete every order
for every other product on the board and leave PMax standing. So this reads,
matches, and updates - it never deletes a row and never creates one.

Matching is on the LINE ITEM id, which is unique, falling back to the order id
when a sheet does not carry one. A line item id that is not on the board yet is
counted and reported rather than silently dropped: that is the export being
ahead of the sync, and it is worth seeing.
"""
from __future__ import annotations

import io
import re
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.orm import Session

from .db import OrderLine

# Which sheet column feeds which stored number. Budget is the same field for
# every product; the spend is per-product, because each one is billed on a
# different basis.
BUDGET_COLS = ("monthly_campaign_budget",)
SPEND_COLS = ("monthly_meta_ad_spend", "monthly_ppc_ad_spend",
              "monthly_linkedin_ad_spend")

MONEY = re.compile(r"[^0-9.\-]")


def _money(v) -> float | None:
    """A number, or None. "$1,215.08" and "-" both arrive from the same sheet."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = MONEY.sub("", str(v).strip())
    if not s or s in ("-", ".", "-."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _rows(src, filename: str = ""):
    """(normalised header -> value) dicts from an xlsx or csv."""
    from .orders_io import normalise_header

    name = str(filename or (src if isinstance(src, (str, Path)) else "")).lower()
    if name.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        data = src if isinstance(src, (str, Path)) else io.BytesIO(src)
        wb = load_workbook(data, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = next(it, None) or []
    else:
        import csv
        if isinstance(src, (str, Path)):
            text = Path(src).read_text(encoding="utf-8-sig", errors="replace")
        else:
            text = src.decode("utf-8-sig", errors="replace")
        it = csv.reader(io.StringIO(text))
        header = next(it, None) or []

    keys: list[str] = []
    for h in header:
        k = normalise_header(h)
        # The export carries two End Dates and two Start Dates on purpose.
        while k in keys:
            k += ".1"
        keys.append(k)
    for row in it:
        yield dict(zip(keys, list(row) + [None] * (len(keys) - len(row))))


def _ids(v) -> set[str]:
    return {x for x in re.split(r"[,\s]+", str(v or "")) if x}


def import_budgets(db: Session, src, filename: str = "") -> dict:
    """Fill in budget and spend from a sheet, touching nothing else."""
    by_line: dict[str, dict] = {}
    by_order: dict[str, dict] = {}
    read = 0
    for r in _rows(src, filename):
        read += 1
        money = {}
        for col in BUDGET_COLS:
            got = _money(r.get(col))
            if got is not None:
                money["budget"] = got
        for col in SPEND_COLS:
            got = _money(r.get(col))
            if got is not None:
                money["spend"] = got
        if not money:
            continue
        line = str(r.get("id") or "").strip()
        order = str(r.get("orders_id") or "").strip()
        # LAST ONE WINS, and that is deliberate. A client running one product
        # across two flights has two rows, and the later flight is the one this
        # month is being paced against.
        if line:
            by_line[line] = money
        if order:
            by_order[order] = money

    matched = order_matched = 0
    for row in db.scalars(select(OrderLine)).all():
        hit = None
        for lid in _ids(row.line_ids):
            if lid in by_line:
                hit = by_line[lid]
                break
        if hit is None:
            for oid in _ids(row.account_ids):
                if oid in by_order:
                    hit = by_order[oid]
                    order_matched += 1
                    break
        else:
            matched += 1
        if hit is None:
            continue
        if "budget" in hit:
            row.budget = hit["budget"]
        if "spend" in hit:
            row.spend = hit["spend"]
    db.commit()
    return {"rows_read": read, "with_money": len(by_line) or len(by_order),
            "matched_on_line_item": matched, "matched_on_order": order_matched,
            "lines_updated": matched + order_matched,
            "not_on_the_board": max(len(by_line) - matched, 0)}
