"""Order-level list: what should arrive, who owns it, and which campaigns just
ended and therefore owe a lifetime report."""
from __future__ import annotations

import csv
import datetime as dt
import io
import re
from pathlib import Path
from dateutil import parser as dp

from sqlalchemy import select
from sqlalchemy.orm import Session

from .db import OrderLine, Report
from .cycle import month_label

COLUMN_ALIASES = {
    "market": {"market", "station", "partner"},
    "client": {"client", "client name", "campaign", "advertiser"},
    "account_ids": {"account", "account id", "account ids", "campaign id", "#"},
    "campaign": {"campaign name", "campaign", "order", "line"},
    "starts_on": {"start", "start date", "campaign start date", "flight start"},
    "ends_on": {"end", "end date", "campaign end date", "flight end", "due"},
    "buyer": {"buyer"},
    "team_member": {"p&a team member", "team member", "pa team member", "owner"},
    "buyer_email": {"buyer email"},
    "team_email": {"team email", "team member email"},
}


def _norm(h: str) -> str:
    return re.sub(r"\s+", " ", (h or "").strip().lower()).strip("*: ")


def _map_headers(headers: list[str]) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, h in enumerate(headers):
        n = _norm(h)
        for field, names in COLUMN_ALIASES.items():
            if n in names and field not in out:
                out[field] = i
    return out


def _date(v):
    v = (v or "").strip()
    if not v:
        return None
    try:
        return dp.parse(v).date()
    except Exception:
        return None


ACC = re.compile(r"\b\d{4,6}\b")


def _rows_from_csv(raw: bytes) -> list[list[str]]:
    text = raw.decode("utf-8-sig", errors="replace")
    return [[(c or "") for c in r] for r in csv.reader(io.StringIO(text))]


def _rows_from_xlsx(raw: bytes, sheet: str | None = None) -> list[list[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    out = []
    for row in ws.iter_rows(values_only=True):
        out.append(["" if v is None else (v.strftime("%Y-%m-%d")
                                          if hasattr(v, "strftime") else str(v))
                    for v in row])
    return out


def import_orders(db: Session, raw, filename: str = "orders.csv",
                  sheet: str | None = None, replace: bool = True,
                  period: str | None = None):
    """Returns the IO export's result dict when it recognizes that format,
    otherwise a plain row count."""
    """Accepts CSV or XLSX, and recognizes the IO tool's own export, which needs
    its own eligibility rules rather than being read as a flat list."""
    from pathlib import Path as _P
    from .orders_io import import_io_export, looks_like_io_export

    blobs = raw if isinstance(raw, list) else [raw]
    if not blobs:
        raise ValueError("No files to import.")

    def header_of(b) -> list[str]:
        """First row, without reading the rest. An export can be 400 MB."""
        if isinstance(b, (str, _P)):
            with open(b, "r", encoding="utf-8-sig", errors="replace", newline="") as fh:
                return next(csv.reader(fh), [])
        if filename.lower().endswith((".xlsx", ".xlsm")):
            rows = _rows_from_xlsx(b, sheet)
            return rows[0] if rows else []
        text = b.decode("utf-8-sig", errors="replace")[:64_000]
        return next(csv.reader(io.StringIO(text)), [])

    # EVERY FILE IS CLASSIFIED ON ITS OWN HEADER.
    #
    # The first version peeked at one file and applied the verdict to all of
    # them. A folder holding one IO export plus anything else - a partner
    # list, a stray sheet, a half-written upload - then went down whichever
    # path that one file chose. Worse, the fallback branch handed file PATHS
    # to a reader expecting bytes, so the failure surfaced as an unrelated
    # error from deep inside the parser with no mention of a file.
    io_exports, others = [], []
    for b in blobs:
        try:
            (io_exports if looks_like_io_export(header_of(b)) else others).append(b)
        except Exception as exc:  # noqa: BLE001 - a bad file names itself
            name = b if isinstance(b, (str, _P)) else filename
            raise ValueError(f"Could not read {Path(str(name)).name}: "
                             f"{type(exc).__name__}: {exc}") from exc

    if io_exports:
        res = import_io_export(db, io_exports, period=period, replace=replace)
        if isinstance(res, dict) and others:
            res["ignored_files"] = [Path(str(o)).name if isinstance(o, (str, _P))
                                    else filename for o in others]
        return res

    # No IO export among them, so treat what is left as a plain list. Read
    # each one as bytes regardless of whether it arrived as a path.
    def as_rows(b) -> list[list[str]]:
        if isinstance(b, (str, _P)):
            return _rows_from_csv(Path(b).read_bytes())
        if filename.lower().endswith((".xlsx", ".xlsm")):
            return _rows_from_xlsx(b, sheet)
        return _rows_from_csv(b)

    all_rows = [r for b in blobs for r in as_rows(b)]
    return _import_rows(db, all_rows, replace=replace)


def import_order_csv(db: Session, raw: bytes, replace: bool = True) -> int:
    return import_orders(db, raw, "orders.csv", replace=replace)


def _import_rows(db: Session, rows: list[list[str]], replace: bool = True) -> int:
    if not rows:
        return 0
    header_idx = 0
    for i, r in enumerate(rows[:10]):
        if _map_headers(r).get("client") is not None or _map_headers(r).get("market") is not None:
            header_idx = i
            break
    cols = _map_headers(rows[header_idx])
    if "client" not in cols:
        raise ValueError("Could not find a client or campaign column in the CSV header.")

    if replace:
        db.query(OrderLine).delete()

    n = 0
    for r in rows[header_idx + 1:]:
        if not any(c.strip() for c in r):
            continue

        def g(field, default=""):
            i = cols.get(field)
            return (r[i].strip() if i is not None and i < len(r) else default)

        client = g("client")
        if not client or client.lower() in {"total", "market"}:
            continue
        accounts = g("account_ids") or " ".join(ACC.findall(client))
        db.add(OrderLine(
            market=g("market"), client=client, account_ids=accounts,
            campaign=g("campaign") or client,
            starts_on=_date(g("starts_on")), ends_on=_date(g("ends_on")),
            buyer=g("buyer"), team_member=g("team_member"),
            buyer_email=g("buyer_email"), team_email=g("team_email"),
        ))
        n += 1
    db.commit()
    return n


def _keyify(client: str, accounts: str) -> set[str]:
    ids = set(ACC.findall(accounts or "")) | set(ACC.findall(client or ""))
    return ids


def _stamp(db: Session, report: Report, ol: OrderLine) -> None:
    report.owner_buyer = ol.buyer_email or ol.buyer
    report.owner_team = ol.team_email or ol.team_member
    if not report.market:
        report.market = ol.market
    _fill_from_roster(db, report)


def _fill_from_roster(db: Session, report: Report) -> None:
    """Fall back to the reporting roster for anything the order line lacks.

    The IO export does not always carry a campaign manager, and it never
    carries the reporting team or the trainer. Those live on the partner.
    """
    from .partners import find as find_partner
    if not report.market:
        return
    p = find_partner(db, report.market)
    if p is None:
        return
    if not report.owner_buyer:
        report.owner_buyer = p.buyer_email or p.buyer
    if not report.owner_team:
        report.owner_team = p.reporting_team


def attach_owners(db: Session, report: Report) -> None:
    """Stamp the market and the owner onto a report from its order lines.

    THE SAME MATCHER THE PRODUCT CHECK USES. This had its own - account ids
    first, then an exact match on the normalized client name - and it missed
    everything the two spellings disagreed on. A report whose market never got
    stamped shows as "no market" on the board and belongs to no partner at all:
    eighty-six of them turned up on one logo. `client_lines` matches on id OR
    name, which is the union that fixed the same problem for products.
    """
    hit = client_lines(db, report.client, report.account_ids) or []
    # A market from an order line, preferring a line that has one - a client
    # can have an order row loaded with a blank business unit.
    for ol in hit:
        if ol.market:
            _stamp(db, report, ol)
            return
    if hit:
        _stamp(db, report, hit[0])
        return
    _fill_from_roster(db, report)


def _overlaps(line, start, end) -> bool:
    """Did this line item deliver at any point inside [start, end]?"""
    s = line.starts_on
    e = line.ends_on
    if end and s and s > end:
        return False
    if start and e and e < start:
        return False
    return True


def is_mapped(product: str) -> bool:
    """A product the tool can actually judge a report against.

    An order line whose product name the map has never seen is kept now rather
    than dropped - the client belongs on the board either way - but it cannot
    be EXPECTED on a report, because nothing knows what "Website Visitor ID"
    looks like when it is there. Failing a report for not carrying it would be
    the tool blaming somebody for a gap in its own dictionary.
    """
    from .checks.products import map_order_product, on_a_report
    # AND SOME MAPPED PRODUCTS ARE STILL NEVER ON A REPORT. Website Visitor ID
    # and Additional Billing are invoiced line items with no widget and never
    # will have one, so expecting them fails every report they are on.
    return (bool(product) and map_order_product(product) is not None
            and on_a_report(product))


def expected_products(db: Session, client: str, account_ids: str,
                     period: str | None = None,
                     lifetime: bool = False,
                     window: tuple | None = None) -> set[str] | None:
    """Products the client's qualifying orders say belong on this report.

    Returns None when the client is not on the order list, so the check stays
    quiet rather than guessing.

    The period filter is not optional in practice. A client with fourteen line
    items across six years has products that stopped running in 2024 sitting
    beside one that is still live, and without a date test every one of them is
    "expected" - which is how a July 2026 report got failed for missing Mobile
    Conquesting whose last line ended on New Year's Eve 2025.
    """
    hit = client_lines(db, client, account_ids)
    if hit is None:
        return None
    # A LIFETIME IS ABOUT THE WHOLE CAMPAIGN, so every product the client ever
    # bought on it belongs on the report - including the ones that stopped
    # months ago and the ones that are paused now. Judged against this month,
    # Burt Young Sales' lifetime was failed for carrying CTV and Native
    # Display, which is exactly what a campaign-to-date report should carry.
    if lifetime:
        # THE CAMPAIGN THIS REPORT COVERS, not everything the client has ever
        # bought. Field Of Dreams' lifetime runs Dec 17 to Jul 13; their
        # Display order starts on 28 July and belongs to the next campaign, so
        # a report without it is not missing anything.
        if window and (window[0] or window[1]):
            hit = [l for l in hit if _overlaps(l, window[0], window[1])]
        # A CANCELLED BUY IS STILL ON THE LIFETIME. Canceling does not mean it
        # never ran - it ran and was stopped - and the lifetime is the report
        # that closes the campaign out, so what it delivered belongs on it.
        return {l.product for l in hit if is_mapped(l.product)}
    if period:
        # An empty result here is not the same as no order list. If every one
        # of a client's products stopped before the period, the honest answer
        # is "nothing was owed" - an empty set, which the check reads as a
        # pass - not None, which it reads as "we cannot say".
        hit = [l for l in hit if _ran_during(l, period)]
    # A paused buy is not delivering, so it is not owed on the report. Nor is a
    # canceled one - which live=False already covers, but says so out loud.
    return {l.product for l in hit if is_mapped(l.product)
            and getattr(l, "live", True)
            and not getattr(l, "canceled", False)}


def quiet_products(db: Session, client: str, account_ids: str,
                   period: str | None = None,
                   lifetime: bool = False) -> set[str]:
    """Products the client has bought but that are not owed on this report.

    Paused line items, and anything whose flight does not touch the month. It
    is not expected - but it is not a surprise on the report either, which is
    the half that was missing: a paused Meta line was failed for BOTH, once for
    being absent and, on another client, once for being present.
    """
    if lifetime:
        return set()          # on a lifetime nothing the client bought is a surprise
    hit = client_lines(db, client, account_ids) or []
    out = set()
    for l in hit:
        if not l.product:
            continue
        # A CANCELED BUY IS QUIET WHATEVER ITS DATES SAY. Roto Rooter's PPC was
        # canceled on 28 July, mid-month, so it ran for most of the period and
        # the date test alone would have left it expected - and then failed the
        # report for showing a product nobody ordered.
        if getattr(l, "canceled", False) or not getattr(l, "live", True):
            out.add(l.product)
        elif period and not _ran_during(l, period):
            out.add(l.product)
    return out


def budgets_for(db: Session, client: str, account_ids: str,
                period: str | None = None) -> dict:
    """What each of this client's live products should spend in the month.

    Summed across the line items that were actually running, because a client
    running one product on two flights is buying both. A line with no budget
    loaded contributes nothing rather than zero - "no budget on file" and "a
    budget of nothing" are different claims and only one of them is true.
    """
    hit = client_lines(db, client, account_ids) or []
    if period:
        hit = [l for l in hit if _ran_during(l, period)]
    out: dict[str, float] = {}
    for l in hit:
        if not l.product or not getattr(l, "live", True):
            continue
        if l.budget is None:
            continue
        out[l.product] = out.get(l.product, 0.0) + float(l.budget)
    return out


def _live_lines(line, period: str | None) -> dict | None:
    """This row's money with the CANCELLED line items taken out.

    {"budget": float|None, "impressions": float|None, "any": bool}, or None
    when the row carries no line-item detail - written before build 88 - in
    which case there is nothing to take out and the merged figures stand.

    None stays None inside it: "the order does not say" is not "the order says
    nothing", and a pacing line that treats the two the same reads 100% under
    on a column that is simply absent.
    """
    detail = getattr(line, "detail", None)
    if not detail:
        return None
    out = {"budget": None, "impressions": None, "any": False}
    for d in detail:
        if not isinstance(d, dict) or d.get("canceled"):
            continue
        if period and not _ran_during(_Window(d.get("starts"), d.get("ends")),
                                     period):
            continue
        out["any"] = True
        for key in ("budget", "impressions"):
            v = d.get(key)
            if v is not None:
                out[key] = float(v) if out[key] is None else out[key] + float(v)
    return out


class _Window:
    """One line item's own flight, shaped like an order line for _ran_during."""

    def __init__(self, starts, ends):
        self.flights = [[starts, ends]]
        self.starts_on, self.ends_on = starts, ends


def ordered_for(db: Session, client: str, account_ids: str,
                period: str | None = None, lifetime: bool = False,
                window: tuple | None = None) -> dict:
    """What each product was bought to do - this month, or in total.

    {product: {"budget": float|None, "impressions": float|None}}, summed over
    the line items behind it. None stays None: "the order does not say" is not
    "the order says nothing", and a pacing line that treats the two the same
    reads 100% under on a column that is simply absent.

    A LIFETIME IS MEASURED AGAINST THE WHOLE CAMPAIGN. Its report covers every
    month the campaign ran, so comparing it to one month's budget says the
    client is nine hundred percent over. It also counts every line item, live
    or not - a campaign that finished is exactly what a lifetime reports on.
    """
    hit = client_lines(db, client, account_ids) or []
    if period and not lifetime:
        hit = [l for l in hit if _ran_during(l, period)]
    # A LIFETIME REPORTS ON ONE CAMPAIGN, NOT ON EVERY ORDER THE CLIENT HAS.
    #
    # Field Of Dreams' lifetime covers Mobile Conquesting, 17 Dec to 13 Jul.
    # A Display order starting 28 Jul - after that campaign finished - was
    # counted into the same goal, so the panel asked a six-page report about
    # 750,000 impressions it was never going to carry and called the whole
    # thing 41% short.
    if window and window[0]:
        hit = [l for l in hit if _overlaps(l, window[0], window[1])]
    out: dict[str, dict] = {}
    # AND ONE GOAL PER LINE ITEM, COUNTED ONCE.
    #
    # "CTV + Video Ads" is one line item sold as two products, so the import
    # writes two order rows for it - both carrying the SAME monthly goal,
    # because it is one goal. Grouped back into one pacing row and then added,
    # it came out doubled: Russell Law's lifetime was measured against 250,000
    # impressions on a campaign sold 125,000, and finished "45% under".
    counted: set[tuple] = set()
    for l in hit:
        if not l.product:
            continue
        if not lifetime and not getattr(l, "live", True):
            continue
        # ONE ROW PER BUY, NOT PER PRODUCT. "CTV + Video Ads" is a single line
        # item with a single goal; pacing each half against the whole goal said
        # CTV was 46% short while Video had nothing to compare against at all.
        name = getattr(l, "sold_with", "") or l.product
        row = out.setdefault(name, {"budget": None, "impressions": None,
                                    "basis": "", "started": None, "days": None,
                                    # A CANCELLED BUY IS NOT SHORT OF ITS GOAL.
                                    # Canceling changes the deal - what it was
                                    # sold to deliver stopped being what it was
                                    # asked to deliver on the day somebody
                                    # stopped it. Sorge's cancelled Meta line
                                    # read "100% short, 503 served against
                                    # 400,000 ordered (20 months at the monthly
                                    # figure on the order)", which is a
                                    # comparison against a campaign that was
                                    # called off.
                                    "stopped": False})
        if getattr(l, "canceled", False) or getattr(l, "complete", False):
            row["stopped"] = True
        # The line item ids behind this row. Two order rows carrying the same
        # ids are two halves of one buy, and its figures belong to it once.
        stamp = (name, getattr(l, "line_ids", "") or "", l.account_ids or "")
        if stamp in counted:
            continue
        counted.add(stamp)
        if not lifetime:
            # WHEN IT LAUNCHED, AND HOW MUCH OF THE MONTH IT HAD.
            #
            # A line that went live on the 28th cannot deliver a month's goal,
            # and pacing it against one says 99% short about a campaign three
            # days old. The date goes on the finding; the day count decides
            # whether there is a finding at all.
            began, ran = _month_window(l, period)
            if began and (row["started"] is None or began < row["started"]):
                row["started"] = began
            if ran is not None:
                row["days"] = ran if row["days"] is None else max(row["days"], ran)
            # A CANCELLED LINE ITEM IS NOT PART OF WHAT THE MONTH IS OWED.
            #
            # The stored row is one answer per client and product, so Houston
            # Concierge Medicine's Social Mirror is three line items on order
            # 54985 added together: two cancelled at 120,000 each and one live
            # at 100,000. Paced against all 340,000 the report read 84% short
            # of a goal that two thirds of was called off. Against the 100,000
            # still being asked for, it is 47% short - which is a real number
            # about a real buy.
            #
            # Only when the line items are actually known: a row loaded before
            # they were kept falls back to the merged figures, which is the
            # answer this always gave.
            live = _live_lines(l, period)
            got = {"budget": None, "impressions": None}
            if live is not None:
                for key in ("budget", "impressions"):
                    got[key] = live.get(key)
                # And the row is only "stopped" if nothing is left running.
                # One cancelled line beside a live one is not a campaign that
                # was called off, and marking it so silenced the finding on
                # the half that is still delivering.
                if live["any"]:
                    row["stopped"] = False
            else:
                for key in ("budget", "impressions"):
                    v = getattr(l, key, None)
                    got[key] = None if v is None else float(v)
            # A CAMPAIGN TOTAL WITH NO MONTHLY FIGURE BESIDE IT.
            #
            # Kerr-Bilt Trailers' Performance Max carries $20,000 for the whole
            # campaign and nothing per month, so the spend row read "-/- no
            # comparison" while the impressions rows above it were being paced
            # against real monthly goals. Two units on one panel, and the one
            # the client is actually billed on was the blank.
            #
            # The lifetime panel already does this in reverse - a monthly goal
            # multiplied out across the flight - and says so. Same here: the
            # total divided by the months it covers, labeled as derived rather
            # than presented as something the order stated.
            months = _months_of(l)
            for src, key in (("total_budget", "budget"),
                             ("total_impressions", "impressions")):
                if got[key] is not None:
                    continue
                whole = getattr(l, src, None)
                if whole is None or not months:
                    continue
                got[key] = float(whole) / months
                row["basis"] = (f"the campaign total over {months} "
                                f"month{'s' if months != 1 else ''}")
            for key in ("budget", "impressions"):
                if got[key] is not None:
                    row[key] = got[key] if row[key] is None else row[key] + got[key]
            continue

        # A LIFETIME IS THE WHOLE CAMPAIGN, and most orders do not carry a
        # campaign total - they carry a monthly goal and a flight. So when the
        # export has a real total it is used, and otherwise the total is the
        # monthly goal across the months the campaign ran. That is a derived
        # figure and the page says so, rather than quietly presenting it as
        # something the order stated.
        months = _months_of(l)
        for src, key in (("total_budget", "budget"),
                         ("total_impressions", "impressions")):
            v = getattr(l, src, None)
            if v is None:
                monthly = getattr(l, key, None)
                if monthly is None or not months:
                    continue
                v = float(monthly) * months
                row["basis"] = (f"{months} month{'s' if months != 1 else ''} "
                                f"at the monthly figure on the order")
            row[key] = float(v) if row[key] is None else row[key] + float(v)
    return out


def _month_window(line, period: str | None):
    """(the day it started delivering this month, days it ran in the month).

    (None, None) with no period to measure against. The start is the LINE's
    own, clipped to the month - what a reader wants to know is "when did this
    go live", and for a line that began in March and is still running that is
    the first of the month.
    """
    if not period:
        return None, None
    y, m = (int(x) for x in period.split("-"))
    first = dt.date(y, m, 1)
    last = dt.date(y + (m == 12), (m % 12) + 1, 1) - dt.timedelta(days=1)

    best_start, best_days = None, None
    windows = [w for w in (getattr(line, "flights", None) or [])
               if isinstance(w, (list, tuple)) and len(w) == 2] \
        or [(line.starts_on, line.ends_on)]
    for w_start, w_end in windows:
        s_, e_ = _as_date(w_start), _as_date(w_end)
        lo = max(s_, first) if s_ else first
        hi = min(e_, last) if e_ else last
        if hi < lo:
            continue
        days = (hi - lo).days + 1
        if best_days is None or days > best_days:
            best_days, best_start = days, lo
    return best_start, best_days


def _months_of(line) -> int:
    """How many months this order ran, at least one.

    Rounded rather than truncated: a flight of 2026-01-09 to 2026-07-08 is six
    months of delivery, not five and a bit.
    """
    start = getattr(line, "order_starts_on", None) or line.starts_on
    end = getattr(line, "order_ends_on", None) or line.ends_on
    if not start or not end or end < start:
        return 0
    return max(1, round(((end - start).days + 1) / 30.44))


def expected_any(db: Session, client: str, account_ids: str,
                 period: str | None = None) -> list:
    """Expectations this client's orders satisfy with EITHER product.

    "Amazon Premium CTV + Video Ads" is one line item that can deliver all of
    its impressions through either half, so a report showing CTV and no Video
    is a normal Amazon month rather than a missing product.
    """
    from .checks.products import any_of_groups

    hit = client_lines(db, client, account_ids) or []
    if period:
        hit = [l for l in hit if _ran_during(l, period)]
    return any_of_groups([l.campaign for l in hit])


def client_lines(db: Session, client: str, account_ids: str):
    """Every order line belonging to this client. None if it is not on the list.

    EVERY ONE, matched by account id OR by name, not whichever matches first.
    That "if not hit" fallback was the bug behind a whole run of false
    positives: a client whose report carries order 31050 also has a live Live
    Chat order under 31171, and because the id match found something, the name
    match that would have found the second order never ran. The report was then
    failed for a Live Chat "with no live order" that was sitting right there,
    and for two products that had stopped in 2024 - the only orders anyone was
    looking at.
    """
    lines = db.scalars(select(OrderLine)).all()
    if not lines:
        return None
    ids = _keyify(client, account_ids)
    norm = re.sub(r"[^a-z0-9]", "", (client or "").lower())
    hit = []
    for l in lines:
        by_id = bool(ids and (ids & _keyify(l.client, l.account_ids)))
        by_name = bool(norm and re.sub(r"[^a-z0-9]", "", l.client.lower()) == norm)
        if by_id or by_name:
            hit.append(l)
    return hit or None


def expected_why(db: Session, client: str, account_ids: str,
                 period: str | None = None) -> list[tuple[str, str]]:
    """Why each product is or is not expected, as trace rows.

    Three rounds of "this is a false positive" all needed the same thing to
    settle them: which orders were being looked at and what their dates were.
    Reading it off the code took a screenshot, a guess and a deploy each time.
    It goes on the finding instead.
    """
    hit = client_lines(db, client, account_ids)
    if hit is None:
        return [("Orders found for this client", "none - no claim is made")]

    def when(l) -> str:
        """The ONE window that settles it, not every window on the order.

        A client running a product across four overlapping flights printed all
        four - "2024-12-13 to 2026-12-31; 2026-02-06 to 2026-12-31; ..." - which
        is a wall of dates you have to subtract in your head. The question is
        always the same: which flight covers this month, or if none does, how
        close the nearest one came.
        """
        wins = [(_as_date(a), _as_date(b))
                for w in (getattr(l, "flights", None) or [])
                if isinstance(w, (list, tuple)) and len(w) == 2
                for a, b in [w]]
        if not wins:
            wins = [(l.starts_on, l.ends_on)]

        def show(a, b) -> str:
            return f"{a or '?'} to {b or 'open'}"

        if not period:
            a, b = wins[0]
            return show(a, b) + (f" (+{len(wins) - 1} more)" if len(wins) > 1 else "")

        y, m = (int(x) for x in period.split("-"))
        first = dt.date(y, m, 1)
        last = dt.date(y + (m == 12), (m % 12) + 1, 1) - dt.timedelta(days=1)
        covering = [(a, b) for a, b in wins
                    if not (b and b < first) and not (a and a > last)]
        if covering:
            a, b = covering[0]
            extra = len(covering) - 1
            return show(a, b) + (f" (+{extra} more covering {month_label(period)})"
                                if extra else "")
        # Nothing covers the month. The nearest end date is what somebody wants
        # to see - "it stopped in June" answers the question on its own.
        ended = [b for _a, b in wins if b]
        if ended:
            return f"ran to {max(ended)}"
        started = [a for a, _b in wins if a]
        if started:
            return f"starts {min(started)}"
        return show(*wins[0])

    rows: list[tuple[str, str]] = []
    for l in sorted(hit, key=lambda x: (x.product or "", x.account_ids or "")):
        if not getattr(l, "live", True):
            verdict = "paused, so not owed either way"
        elif period and not _ran_during(l, period):
            verdict = f"not running in {month_label(period)}"
        else:
            verdict = "counted"
        rows.append((f"{l.product or 'unmapped'} · order {l.account_ids or '?'}",
                     f"{when(l)} · {verdict}"))
    return rows


def _as_date(v):
    if not v:
        return None
    if isinstance(v, dt.date):
        return v
    try:
        return dt.date.fromisoformat(str(v)[:10])
    except ValueError:
        return None


def _ran_during(line, period: str) -> bool:
    """Did this product actually run in the report's month?

    Asked of each order's OWN window, not of the merged span across them.
    Blair Regional YMCA ran Social Mirror CTV until June 2026 and starts again
    on 1 August. Merged, that is one flight from 2025 to December 2026, and
    July - a month in which no Social Mirror CTV ran at all - sits inside it.
    The July report was failed twice over for it: once for the product, once
    for the CTV publishers widget the product would have owed.
    """
    y, m = (int(x) for x in period.split("-"))
    start = dt.date(y, m, 1)
    end = dt.date(y + (m == 12), (m % 12) + 1, 1) - dt.timedelta(days=1)

    def touches(s, e) -> bool:
        s, e = _as_date(s), _as_date(e)
        if e and e < start:
            return False
        if s and s > end:
            return False
        return True

    windows = [w for w in (getattr(line, "flights", None) or [])
               if isinstance(w, (list, tuple)) and len(w) == 2]
    if windows:
        return any(touches(w[0], w[1]) for w in windows)
    # Nothing recorded - an order line loaded before the windows were kept.
    # Fall back to the merged span, which is what this used to do.
    return touches(line.starts_on, line.ends_on)


def completeness(db: Session, market: str, period: str) -> dict:
    """What should have arrived for this market and period, versus what did."""
    lines = db.scalars(select(OrderLine).where(OrderLine.market == market)).all() if market \
        else db.scalars(select(OrderLine)).all()
    got = db.scalars(select(Report).where(Report.period == period)).all()
    if market:
        got = [r for r in got if (r.market or "") == market]

    got_ids = {i for r in got if not r.is_lifetime for i in _keyify(r.client, r.account_ids)}
    life_ids = {i for r in got if r.is_lifetime for i in _keyify(r.client, r.account_ids)}

    y, m = (int(x) for x in period.split("-"))
    period_end = (dt.date(y + (m == 12), (m % 12) + 1, 1) - dt.timedelta(days=1))
    period_start = dt.date(y, m, 1)

    # "Team member" was a column from the original schema that nothing fills
    # any more - the IO export has no such field, so it rendered empty on every
    # row. The reporter is the person actually wanted here, and that lives on
    # the partner roster.
    from .partners import find as find_partner
    pcache: dict[str, object] = {}

    def reporter_for(market: str) -> str:
        if market not in pcache:
            pcache[market] = find_partner(db, market)
        p = pcache[market]
        return p.reporting_team if p else ""

    missing, lifetime_due = [], []
    for ol in lines:
        ids = _keyify(ol.client, ol.account_ids)
        if ol.starts_on and ol.starts_on > period_end:
            continue
        if ol.ends_on and ol.ends_on < period_start:
            continue
        if not ids or not (ids & got_ids):
            missing.append({"client": ol.client, "accounts": ol.account_ids,
                            "line_ids": ol.line_ids, "market": ol.market,
                            "buyer": ol.buyer, "team": reporter_for(ol.market)})
        if ol.needs_lifetime and ol.ends_on and period_start <= ol.ends_on <= period_end:
            if not (ids & life_ids):
                lifetime_due.append({"client": ol.client, "accounts": ol.account_ids,
                                     "line_ids": ol.line_ids, "market": ol.market,
                                     "ended": ol.ends_on.isoformat(),
                                     "buyer": ol.buyer, "team": reporter_for(ol.market)})
    return {"expected": len(lines), "received": len(got),
            "missing": missing, "lifetime_due": lifetime_due}
